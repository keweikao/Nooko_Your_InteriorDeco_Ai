import base64
import json
import os
import io
import logging

from google.cloud import firestore
from google.cloud import storage
from google.cloud import vision
import pdfplumber
import openpyxl

# --- Client Initialization ---
# It's a best practice to initialize clients outside of the function handler
# to take advantage of connection reuse.
storage_client = storage.Client()
firestore_client = firestore.Client()
vision_client = vision.ImageAnnotatorClient()

# --- Environment Variables ---
PROJECTS_COLLECTION = os.getenv("PROJECTS_COLLECTION", "projects")

# --- Logging ---
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def _get_file_from_gcs(gcs_uri: str) -> tuple[bytes, str]:
    """
    Downloads a file from a GCS URI.

    Args:
        gcs_uri: The GCS URI of the file (e.g., "gs://bucket_name/file_path").

    Returns:
        A tuple containing the file content as bytes and the bucket name.
    """
    try:
        bucket_name, blob_name = gcs_uri.replace("gs://", "").split("/", 1)
        bucket = storage_client.bucket(bucket_name)
        blob = bucket.blob(blob_name)
        file_content = blob.download_as_bytes()
        logger.info(f"Successfully downloaded file from {gcs_uri}")
        return file_content, bucket_name
    except Exception as e:
        logger.error(f"Failed to download file from GCS URI {gcs_uri}: {e}")
        raise


def _parse_pdf(content: bytes) -> str:
    """Extracts text from PDF content."""
    text = ""
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    logger.info(f"Successfully parsed PDF, extracted {len(text)} characters.")
    return text


def _parse_excel(content: bytes) -> str:
    """Extracts text from Excel content."""
    text = ""
    workbook = openpyxl.load_workbook(io.BytesIO(content))
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        text += f"--- Sheet: {sheet_name} ---\n"
        for row in sheet.iter_rows(values_only=True):
            row_text = "\t".join([str(cell) for cell in row if cell is not None])
            text += row_text + "\n"
    logger.info(f"Successfully parsed Excel, extracted {len(text)} characters.")
    return text


def _parse_image_ocr(gcs_uri: str) -> str:
    """Performs OCR on an image file in GCS."""
    image = vision.Image()
    image.source.gcs_image_uri = gcs_uri
    response = vision_client.text_detection(image=image)
    if response.error.message:
        raise Exception(f"Vision API error: {response.error.message}")
    text = response.full_text_annotation.text
    logger.info(f"Successfully performed OCR, extracted {len(text)} characters.")
    return text


def _save_result_to_firestore(project_id: str, extracted_text: str):
    """
    Saves the extracted text and updates the project status in Firestore.
    """
    try:
        doc_ref = firestore_client.collection(PROJECTS_COLLECTION).document(project_id)
        doc_ref.update({
            "quote_analysis_status": "completed",
            "original_quote_content": extracted_text,
            "quote_analysis_completed_at": firestore.SERVER_TIMESTAMP
        })
        logger.info(f"Successfully saved analysis for project {project_id} to Firestore.")
    except Exception as e:
        logger.error(f"Failed to save result to Firestore for project {project_id}: {e}")
        raise


def process_quote(event, context):
    """
    Cloud Function entry point triggered by a Pub/Sub message.

    This function processes an uploaded quote file. It downloads the file from GCS,
    parses its content based on the MIME type, and saves the extracted text
    back to Firestore.

    Args:
        event (dict): The event payload from Pub/Sub. The `data` field is a
                      base64-encoded JSON string containing file metadata.
        context (google.cloud.functions.Context): The event metadata.
    """
    # --- Purpose: Main handler for the background quote processing. ---
    # --- Input: Pub/Sub event from the main API's /upload endpoint. ---
    # --- Output: Writes extracted text to the project's Firestore document. ---
    logger.info(f"Function triggered by messageId: {context.event_id} published at {context.timestamp}")

    try:
        # 1. Parse the incoming Pub/Sub message
        if 'data' not in event:
            logger.error("No 'data' field in the event payload.")
            return

        message_data_str = base64.b64decode(event['data']).decode('utf-8')
        message_data = json.loads(message_data_str)
        logger.info(f"Received message: {message_data}")

        project_id = message_data.get("project_id")
        gcs_uri = message_data.get("gcs_uri")
        content_type = message_data.get("content_type")

        if not all([project_id, gcs_uri, content_type]):
            logger.error(f"Missing required data in Pub/Sub message: {message_data}")
            return

        # 2. Download file from GCS
        file_content, _ = _get_file_from_gcs(gcs_uri)
        extracted_text = ""

        # 3. Select parser based on content type
        if content_type == "application/pdf":
            extracted_text = _parse_pdf(file_content)
        elif "spreadsheetml" in content_type or "ms-excel" in content_type:
            extracted_text = _parse_excel(file_content)
        elif content_type in ["image/jpeg", "image/png"]:
            # Vision API works directly with GCS URI
            extracted_text = _parse_image_ocr(gcs_uri)
        else:
            logger.warning(f"Unsupported content type '{content_type}' for parsing.")
            extracted_text = "Unsupported file type for analysis."

        # 4. Save the result to Firestore
        _save_result_to_firestore(project_id, extracted_text)

        logger.info(f"Successfully processed file for project {project_id}.")

    except Exception as e:
        logger.error(f"Unhandled exception in process_quote: {e}", exc_info=True)
        # Optionally, update Firestore with an error status
        project_id = message_data.get("project_id") if 'message_data' in locals() else None
        if project_id:
            _save_result_to_firestore(project_id, f"Error processing file: {e}")
        # The function will implicitly ACK the message on successful completion
        # or NACK on unhandled exception, causing Pub/Sub to retry.
        return
